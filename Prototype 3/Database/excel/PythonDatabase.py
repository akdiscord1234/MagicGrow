# plan-
# IN TEMPERATURESHEET,
# COlUMN A1 is temperature
# COLUMN A2 is time
# SAME FOR EC, pH, and WATERPUMP UP & WATERPUMP DOWN
#
# STEPS-
# 1. IMPORT + SETUP EVERYTHING FOR READING VALUES, INLCUDING KASA, PHIDGET, AND ARDUINO
# 2. IN LOOP, SAVE EACH VALUE (LIKE TEMPERATURE) TO EACH RESPECTIVE SHEET (LIKE TEMPERATURE SHEET) AND TIME
# 3. AUTOSAVE FILE EVERY 10 MINUTES

from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import random
import time
import re

import time
import asyncio
from kasa import *
from Phidget22.Devices.TemperatureSensor import *
from Phidget22.Devices.PHSensor import *
from time import sleep
from pantry_wrapper import *
from datetime import *
import pandas as pd

# EC Sensor Setup

from pyfirmata2 import Arduino, util
import time

# Define constants
TdsSensorPin = 0
VREF = 5.0
SCOUNT = 30

# Initialize Arduino board
board = Arduino('/dev/serial/by-id/usb-Arduino__www.arduino.cc__0043_85733323939351A04291-if00')  # Replace with the correct port for your Arduino

analogBuffer = [0] * SCOUNT
analogBufferTemp = [0] * SCOUNT
analogBufferIndex = 0
copyIndex = 0

averageVoltage = 0
tdsValue = 0
temperature = 19.5

it = util.Iterator(board)
it.start()

# Configure analog pin
tds_sensor = board.get_pin('a:' + str(TdsSensorPin) + ':i')
time.sleep(1)

def get_median_num(b_array, i_filter_len):
    b_tab = b_array.copy()
    for i in range(i_filter_len - 1):
        for j in range(i_filter_len - i - 1):
            if b_tab[j] > b_tab[j + 1]:
                b_temp = b_tab[j]
                b_tab[j] = b_tab[j + 1]
                b_tab[j + 1] = b_temp

    if i_filter_len & 1:
        b_temp = b_tab[(i_filter_len - 1) // 2]
    else:
        b_temp = (b_tab[i_filter_len // 2] + b_tab[i_filter_len // 2 - 1]) // 2

    return b_temp


#temperature sensing
temp = TemperatureSensor()
temp.setDeviceSerialNumber(702002)
temp.setHubPort(0)
temp.openWaitForAttachment(5000)

#pH sensing
pH = PHSensor()
pH.setDeviceSerialNumber(702002)
pH.setHubPort(1)
pH.openWaitForAttachment(5000)

pantry_id = "759b6bd3-6956-480d-afa4-5569177af40b"

#main get_contents
temperaturePantry = get_contents(pantry_id, "temperature", return_type="body")
pHPantry = get_contents(pantry_id, "pH", return_type="body")
ECPantry = get_contents(pantry_id, "EC", return_type="body")

execution = "none"

async def main(execution):

    #Discovering Plugs
    WaterPumpDown = SmartPlug("192.168.4.99")
    await WaterPumpDown.update()

    WaterPumpUp = SmartPlug("192.168.4.100")
    await WaterPumpUp.update()

    ECPumpUp = SmartPlug("192.168.4.101")
    await ECPumpUp.update()

    pHPumpDown = SmartPlug("192.168.4.87")
    await pHPumpDown.update()

    pHPumpUp = SmartPlug("192.168.4.98")
    await pHPumpUp.update()

    LightPlug = SmartPlug("192.168.4.89")
    await LightPlug.update()

    #functions

    if execution == "none":
        pass

    #Water Pump Up
    if execution == "WaterPumpUpOn":
        await WaterPumpUp.turn_on()

    if execution == "WaterPumpUpOff":
        await WaterPumpUp.turn_off()


    #Water Pump Down
    if execution == "WaterPumpDownOn":
        await WaterPumpDown.turn_on()

    if execution == "WaterPumpDownOff":
        await WaterPumpDown.turn_off()

    #EC Pump Up
    if execution == "ECPumpUpOn":
        await ECPumpUp.turn_on()

    if execution == "ECPumpUpOff":
        await ECPumpUp.turn_off()

    #pH Pump Up
    if execution == "pHPumpUpOn":
        await pHPumpUp.turn_on()

    if execution == "pHPumpUpOff":
        await pHPumpUp.turn_off()

    #pH Pump Down
    if execution == "pHPumpDownOn":
        await pHPumpDown.turn_on()

    if execution == "pHPumpDownOff":
        await pHPumpDown.turn_off()

    #Light Plug
    if execution == "LightOn":
        await LightPlug.turn_on()

    if execution == "LightOff":
        await LightPlug.turn_off()


# while True:
#     sleep(5)

#     asyncio.get_event_loop().run_until_complete(main(execution))

#     # add nutrients get_contents
#     WaterUpPumpChanges = get_contents(pantry_id, "WaterUpPumpChanges", return_type="body")
#     WaterDownPumpChanges = get_contents(pantry_id, "WaterDownPumpChanges.json", return_type="body")
#     pHUpChanges = get_contents(pantry_id, "pHUpChanges", return_type="body")
#     pHDownChanges = get_contents(pantry_id, "pHDownChanges", return_type="body")
#     ECChanges = get_contents(pantry_id, "ECChanges", return_type="body")
#     LightChanges = get_contents(pantry_id, "LightChanges", return_type="body")
#     TemperatureChanges = get_contents(pantry_id, "TemperatureChanges", return_type="body")


#     #list of Changes
#     WaterDownPumpChangesList = [value for value in WaterDownPumpChanges.values() if not isinstance(value, (int, float))]
#     print(WaterDownPumpChangesList)

#     WaterUpPumpChangesList = [value for value in WaterUpPumpChanges.values() if not isinstance(value, (int, float))]
#     print(WaterUpPumpChangesList)

#     pHUpChangesList = [value for value in pHUpChanges.values() if not isinstance(value, (int, float))]
#     print(pHUpChangesList)

#     pHDownChangesList = [value for value in pHDownChanges.values() if not isinstance(value, (int, float))]
#     print(pHDownChangesList)

#     ECChangesList = [value for value in ECChangesList.values() if not isinstance(value, (int, float))]
#     print(ECChangesList)

#     LightChangesList = [value for value in LightChangesList.values() if not isinstance(value, (int, float))]
#     print(LightChangesList)

#     TemperatureChangesList = [value for value in TemperatureChangesList.values() if not isinstance(value, (int, float))]
#     print(TemperatureChangesList)

#     #Pump Changes

#     #Water Down Pump Changes
#     if WaterDownPumpChangesList[-1] == "True":
#         execution = "WaterPumpDownOn"
#     elif WaterDownPumpChangesList[-1] == "False":
#         execution = "WaterPumpDownOff"

#     # Water Up Pump Changes
#     if WaterUpPumpChangesList[-1] == "True":
#         execution = "WaterPumpUpOn"
#     if WaterUpPumpChangesList[-1] == "False":
#         execution = "WaterPumpUpOff"

#     # pH Up Pump Changes
#     if pHUpChangesList[-1] == "True":
#         execution = "pHPumpUpOn"
#     if pHUpChangesList[-1] == "False":
#         execution = "pHPumpUpOff"

#     # pH Down Pump Changes
#     if pHDownChangesList[-1] == "True":
#         execution = "pHPumpDownOn"
#     if pHDownChangesList[-1] == "False":
#         execution = "pHPumpDownOff"

#     # EC Pump Changes
#     if ECChangesList[-1] == "True":
#         execution = "ECPumpUpOn"
#     if ECChangesList[-1] == "False":
#         execution = "ECPumpUpOff"

#     #Other Changes

#     # Light Changes
#     if LightChangesList[-1] == "True":
#         execution = "LightOn"

#     if LightChangesList[-1] == "False":
#         execution = "LightOff"

#     #EC code

#         # Read analog value
#         analog_value = tds_sensor.read()

#         # Store analog value into the buffer
#         if analog_value is not None:
#             analogBuffer[analogBufferIndex] = analog_value * 1024
#         else:
#             analogBuffer[analogBufferIndex] = 0
#         analogBufferIndex += 1

#         if analogBufferIndex == SCOUNT:
#             analogBufferIndex = 0

#         # Read analog value every 40 milliseconds
#         time.sleep(0.04)

#         # Calculate median value using the median filtering algorithm
#         for copyIndex in range(SCOUNT):
#             analogBufferTemp[copyIndex] = analogBuffer[copyIndex]

#         median_value = get_median_num(analogBufferTemp, SCOUNT)

#         # Convert to voltage value
#         averageVoltage = median_value * VREF / 1024.0

#         # Your additional logic for TDS and temperature calculation can be added here
#         finalEC = ((133.42 * (averageVoltage ** 3)) - 255.86 * (averageVoltage ** 2) + 857.39 * averageVoltage) * 0.5


#     #main append
#     temperaturePantry[str(datetime.now())] = temp.getTemperature()
#     pHPantry[str(datetime.now())] = pH.getPH()
#     ECPantry[str(datetime.now())] = finalEC

#     append_basket(pantry_id, "temperature", temperaturePantry, return_type="body") #temperature
#     append_basket(pantry_id, "pH", pHPantry, return_type="body")  # pH
#     append_basket(pantry_id, "EC", ECPantry, return_type="body")  # EC

# ---------------------------------------------------------------- -------------------------------- 

def current_time():
    char_to_remove = "'"
    
    # Replace the specified character with an empty string
    
    # local_time_A = local_time.translate({ord("'"):None})
    
    # local_time_A = local_time.replace("'", "", 1)
    local_time_A = local_time.replace("''", '')
    # Use regular expressions to extract the function name and arguments
    # pattern = r"def\s+([a-zA-Z0-9_]+)\s*\((.*?)\)\s*:"
    # match = re.match(pattern, new_func_str)
    # if match:
    #     func_name = match.group(1)
    #     args = match.group(2)
    #     new_func_str = f"def {func_name}({args}):" + new_func_str[match.end():]

    
    timee = f"{local_time_A}"
    
    return timee



#Create a workbook object
#wb = Workbook()

#load existing spreadsheet
wb = load_workbook(r"C:\Users\kavet\OneDrive\Documents\GitHub\MagicGrow\Prototype 3\Database\excel\Database.xlsx")

#Create Worksheets
ws_Temperature = wb["Temperature"] #THE TEMPERATURE SHEET IS ALSO THE LIGHT SHEET BECAUSE TEMPERATURE MODIFIES LIGHT
ws_EC = wb["EC"]
ws_pH = wb["pH"]
ws_WaterPumpUp = wb["WaterPumpUp"]
ws_WaterPumpDown = wb["WaterPumpDown"]

#Columns

temp_column_a = ws_Temperature["a"]
temp_column_c = ws_Temperature["c"]

# Set a variable
# name = ws["A3"].value
# color = ws["B3"].value

# Print something from our worksheet
# print(f'{name} : {color}')

# Grab a whole column



# #For loop
# for cell1 in temp_column_a:
#     print(cell1.value)
# for cell2 in temp_column_c:
#     print(cell2.value)

# #Grab a range
# range = ws_Temperature["A1":"D30"]

# #Loop
# for cell in range: #with tuple
#     for x in cell: #removes tuple parenthesis
#         print(x.value)

# #Getting the number of cells in a column

# # Specify the column number (e.g., column A is 1, B is 2, etc.)
# temp_column_a_number = 1

# # Get the column letter corresponding to the column number
# column_letter = chr(64 + temp_column_a_number)

# # List to store non-empty cell values in the specified column
# temp_column_a_cells = ws_Temperature[column_letter]

# # Calculate number of non-empty cells in the column
# num_of_in_temp_column_a = sum(1 for cell in temp_column_a_cells if cell.value is not None)

# # Print the number of cells in the column
# print(f"Number of non-empty cells in column {column_letter}: {num_of_in_temp_column_a}")

#To set value of a cell

# c2 = sheet.cell(row= 1 , column = 2) 
# c2.value = "RAI"

def save_file(): #function for saving file
    wb.save(r"C:\Users\kavet\OneDrive\Documents\GitHub\MagicGrow\Prototype 3\Database\excel\Database.xlsx")


#Function for reading number of values in a cell
def number_of_values_in_cell(column_number_S, worksheeT): # First specify column number (e.g., column A is 1, B is 2, etc.), then worksheet
        
    # Specify the column number (e.g., column A is 1, B is 2, etc.)
    column_number = column_number_S

    # Get the column letter corresponding to the column number
    column_letter = chr(64 + column_number)

    # List to store non-empty cell values in the specified column
    column_cells = worksheeT[column_letter]

    # Calculate number of non-empty cells in the column
    num_of_in_column = sum(1 for cell in column_cells if cell.value is not None)
    
    return num_of_in_column

# def append_new_cell(append_value, worksheet, columN):
#     cValue = worksheet.cell(row=int(number_of_values_in_cell(1, worksheet)) + 1, column = columN)
#     cValue.value = append_value
#     save_file()

def append_new_cell(append_value, ws, col_num, row_num):
    cValue = ws.cell(row=row_num, column=col_num) # Specify cell
    cValue.value = str(append_value)# Ensure append_value is a string

def append_row_Temperature(New_Temperature, New_Light_On, New_Meeting_Requirements, New_Time, row_nuM):
    append_new_cell(New_Temperature, ws_Temperature, 1, row_nuM)
    append_new_cell(New_Light_On, ws_Temperature, 2, row_nuM)
    append_new_cell(New_Time, ws_Temperature, 3, row_nuM)
    append_new_cell(New_Meeting_Requirements, ws_Temperature, 4, row_nuM)
    save_file()

#Temperature Add New Cell
# New_Temperature = 29
# New_Light_On = "ON"
# New_Meeting_Requirements = "ON"
# New_Time = "02:00:00 PM"

while True:
    #Append new row in Temperature
    New_Temperature = [random.randint(20, 40)]
    New_Light_OnOff = ["ON", "OFF"]
    Temperature_New_Meeting_Requirements = ["YES", "NO"]
    New_Light_OnOff = random.choice(New_Light_OnOff)
    Temperature_New_Meeting_Requirements = random.choice(Temperature_New_Meeting_Requirements)
    #Time
    time_object = time.localtime()
    time_object = time.gmtime()
    local_time = time.strftime('%B %d %Y %H:%M:%S', time_object)
    
    append_row_Temperature(New_Temperature[0], New_Light_OnOff, Temperature_New_Meeting_Requirements, local_time[:], number_of_values_in_cell(1, ws_Temperature))
    
    #Append new row in EC
    
    New_EC = random.randint(3, 5)
    BaseABUsed = ["ON", "OFF"]
    BaseABUsed = random.choice(BaseABUsed)
    EC_New_Meeting_Requirements = ["YES", "NO"]
    EC_New_Meeting_Requirements = random.choice(EC_New_Meeting_Requirements)
    
    append_new_cell(New_EC, ws_EC, 1, number_of_values_in_cell(1, ws_EC))
    append_new_cell(BaseABUsed, ws_EC, 2, number_of_values_in_cell(1, ws_EC))
    append_new_cell(local_time, ws_EC, 3, number_of_values_in_cell(1, ws_EC))
    append_new_cell(EC_New_Meeting_Requirements, ws_EC, 4, number_of_values_in_cell(1, ws_EC))
    save_file()
    
    #Append new row in pH
    print()
    New_pH = random.randint(3, 5)
    pHUpPumpUsed = ["ON", "OFF"]
    pHUpPumpUsed = random.choice(pHUpPumpUsed)
    pHDownPumpUsed = ["ON", "OFF"]
    pHDownPumpUsed = random.choice(pHDownPumpUsed)
    pH_New_Meeting_Requirements = ["YES", "NO"]
    pH_New_Meeting_Requirements = random.choice(pH_New_Meeting_Requirements)
    
    append_new_cell(New_EC, ws_EC, 1, number_of_values_in_cell(1, ws_EC))
    append_new_cell(BaseABUsed, ws_EC, 2, number_of_values_in_cell(1, ws_EC))
    append_new_cell(local_time, ws_EC, 3, number_of_values_in_cell(1, ws_EC))
    append_new_cell(EC_New_Meeting_Requirements, ws_EC, 4, number_of_values_in_cell(1, ws_EC))
    save_file()
    
    
    
    exit()