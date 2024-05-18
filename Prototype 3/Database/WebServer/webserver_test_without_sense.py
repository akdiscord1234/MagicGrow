
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import random
import time
import re

import time
from time import sleep
from pantry_wrapper import *
from datetime import *
import pandas as pd

# EC Sensor Setup

import time

#temperature sensing
temperature = 29

#pH sensing
pH = 6

#Excel Functions


#load existing spreadsheet
wb = load_workbook(r"C:\Users\kavet\OneDrive\Documents\GitHub\MagicGrow\Prototype 3\Database\excel\Database.xlsx")

#Create Worksheets
ws_Temperature = wb["Temperature"] #THE TEMPERATURE SHEET IS ALSO THE LIGHT SHEET BECAUSE TEMPERATURE MODIFIES LIGHT
ws_EC = wb["EC"]
ws_pH = wb["pH"]
ws_WaterPumpUp = wb["WaterPumpUp"]
ws_WaterPumpDown = wb["WaterPumpDown"]

def save_file(): #function for saving file
    wb.save(r"C:\Users\kavet\OneDrive\Documents\GitHub\MagicGrow\Prototype 3\Database\excel\Database.xlsx")


#Function for reading number of values in a cell
def number_of_values_in_cell(column_number_S, worksheeT): # First specify column number (e.g., column A is 1, B is 2, etc.), then worksheet
        
    # Specify the column number (e.g., column A is 1, B is 2, etc.)
    column_number = column_number_S

    # Get the column letter corresponding to the column number
    column_letter = chr(64 + column_number)

    # List to store non-empty cell values in the specified column
    column_cells = worksheeT[column_letter]

    # Calculate number of non-empty cells in the column
    num_of_in_column = sum(1 for cell in column_cells if cell.value is not None)
    
    return num_of_in_column

# def append_new_cell(append_value, worksheet, columN):
#     cValue = worksheet.cell(row=int(number_of_values_in_cell(1, worksheet)) + 1, column = columN)
#     cValue.value = append_value
#     save_file()

def append_new_cell(append_value, ws, col_num, row_num):
    cValue = ws.cell(row=row_num, column=col_num) # Specify cell
    cValue.value = str(append_value)# Ensure append_value is a string

def append_row_Temperature(New_Temperature, New_Light_On, New_Meeting_Requirements, New_Time, row_nuM):
    append_new_cell(New_Temperature, ws_Temperature, 1, row_nuM)
    append_new_cell(New_Light_On, ws_Temperature, 2, row_nuM)
    append_new_cell(New_Time, ws_Temperature, 3, row_nuM)
    append_new_cell(New_Meeting_Requirements, ws_Temperature, 4, row_nuM)
    save_file()


while True:
    
    finalEC = 3.3
    
    #Recording these values to the spreadsheet code HERE
    #Append new row in Temperature
    New_Temperature = [random.randint(20, 40)]
    New_Light_OnOff = ["ON", "OFF"]
    Temperature_New_Meeting_Requirements = ["YES", "NO"]
    New_Light_OnOff = random.choice(New_Light_OnOff)
    Temperature_New_Meeting_Requirements = random.choice(Temperature_New_Meeting_Requirements)
    #Time
    time_object = time.localtime()
    time_object = time.gmtime()
    local_time = time.strftime('%B %d %Y %H:%M:%S', time_object)
    
    append_row_Temperature(New_Temperature[0], New_Light_OnOff, Temperature_New_Meeting_Requirements, local_time[:], number_of_values_in_cell(1, ws_Temperature))
    
    #Append new row in EC
    
    New_EC = random.randint(3, 5)
    BaseABUsed = ["ON", "OFF"]
    BaseABUsed = random.choice(BaseABUsed)
    EC_New_Meeting_Requirements = ["YES", "NO"]
    EC_New_Meeting_Requirements = random.choice(EC_New_Meeting_Requirements)
    
    append_new_cell(New_EC, ws_EC, 1, number_of_values_in_cell(1, ws_EC))
    append_new_cell(BaseABUsed, ws_EC, 2, number_of_values_in_cell(1, ws_EC))
    append_new_cell(local_time, ws_EC, 3, number_of_values_in_cell(1, ws_EC))
    append_new_cell(EC_New_Meeting_Requirements, ws_EC, 4, number_of_values_in_cell(1, ws_EC))
    save_file()
    
    #Append new row in pH
    print()
    New_pH = random.randint(3, 5)
    pHUpPumpUsed = ["ON", "OFF"]
    pHUpPumpUsed = random.choice(pHUpPumpUsed)
    pHDownPumpUsed = ["ON", "OFF"]
    pHDownPumpUsed = random.choice(pHDownPumpUsed)
    pH_New_Meeting_Requirements = ["YES", "NO"]
    pH_New_Meeting_Requirements = random.choice(pH_New_Meeting_Requirements)
    
    append_new_cell(New_EC, ws_EC, 1, number_of_values_in_cell(1, ws_EC))
    append_new_cell(BaseABUsed, ws_EC, 2, number_of_values_in_cell(1, ws_EC))
    append_new_cell(local_time, ws_EC, 3, number_of_values_in_cell(1, ws_EC))
    append_new_cell(EC_New_Meeting_Requirements, ws_EC, 4, number_of_values_in_cell(1, ws_EC))
    save_file()
    
    #do that code....

    #functions for controling the plugs

    def WaterPumpControls(control_name):
        #Water Pump Down controls
        if control_name == "down_on":
            print("Water Pump Down on")
        elif control_name == "down_off":
            print("Water Pump Down off")
        elif control_name == "down_toggle":
            print("Water Pump Down toggle")
        #Water Pump Up controls
        if control_name == "up_on":
            print("Water Pump Up on")
        elif control_name == "up_off":
            print("Water Pump Up off")
        elif control_name == "up_toggle":
            print("Water Pump Up toggle")

    def pHPumpControls(control_name):
        #pH Pump Down controls
        if control_name == "on":
            print("pH Pump Down on")
        elif control_name == "off":
            print("pH Pump Down off")
        elif control_name == "toggle":
            print("pH Pump Down toggle")

        #pH Pump Up controls
        if control_name == "on":
            print("pH Pump Up on")
        elif control_name == "off":
            print("pH Pump Up off")
        elif control_name == "toggle":
            print("pH Pump Up toggle")

    def ECPumpControls(control_name):
        #EC Pump Up controls
        
        if control_name == "on":
            print("EC Pump on")
        elif control_name == "off":
            print("EC Pump off")
        elif control_name == "toggle":
            print("Toggle EC")
        
        

    def LightPlugControls(control_name):
        if control_name == "on":
            print("Light on")
        elif control_name == "off":
            print("Light off")
        elif control_name == "toggle":
            print("Light toggle")

    def get_current_value(value_name):
        if value_name == "temperature":
            return temperature
        if value_name == "pH":
            return pH
        if value_name == "EC":
            return finalEC